import asyncio
import time
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import openpyxl
import random
import logging
from datetime import datetime
import os
import copy
# Definindo as variáveis de configuração
intervalo_min = 3.5  # Exemplo: 3.5 segundos
intervalo_max = 7.5
EMAIL_HOST = "smtp.hostinger.com"
EMAIL_PORT = 587
EMAIL_USER = "no-reply@worldopportunity.com.br"
EMAIL_PASSWORD = "#Curso123456"
FROM_EMAIL = "no-reply@worldopportunity.com.br"
LOG_FILE       = "email_log.txt"

LIMITE_DIARIO = 100  # Número máximo de e-mails a serem enviados por dia
corpo_primeiro_email_padrao=''
corpo_segundo_email_padrao=''
corpo_terceiro_email_padrao=''
emails_enviados =[]
alterou_tabela=False
def altera_tabela(row_index,column_index,value,sheet):
    global alterou_tabela
    sheet.cell(row=row_index, column=column_index, value=value)
    alterou_tabela = True
    
##sheet.cell(row=row_index, column=indice_primeiro_email, value=time.strftime('%Y-%m-%d %H:%M:%S'))
# Contador de e-mails enviados no dia atual
contador_emails_enviados = 0

# Configuração do log para registrar erros
logging.basicConfig(filename=LOG_FILE, level=logging.INFO, format='%(asctime)s - %(message)s')

# Função para carregar a planilha com verificação de erro

def carregar_planilha():
    nome_planilha = 'PROSPECÇÃO ECONODATA.xlsx'
    try:
        if not os.path.exists(nome_planilha):
            raise FileNotFoundError("O arquivo " + nome_planilha + " não foi encontrado.")
        workbook = openpyxl.load_workbook(nome_planilha)
        sheet = workbook.active
        return sheet, workbook
    except FileNotFoundError as e:
        logging.error(f"Erro ao carregar o arquivo: {str(e)}")
        print(str(e))
        return None, None  # Retorna None para indicar falha, mas não encerra o programa
    except Exception as e:
        logging.error(f"Erro ao carregar a planilha: {str(e)}")
        print(f"Erro ao carregar a planilha: {str(e)}")
        return None, None  # Retorna None para indicar falha, mas não encerra o programa



# Função auxiliar para calcular os dias passados desde a data fornecida
##def dias_passados(data_str):
##    try:
##        data_email = datetime.strptime(data_str, '%Y-%m-%d %H:%M:%S')  # Converte string de data para objeto datetime
##        data_atual = datetime.now()  # Obtém a data e hora atuais
##        delta = data_atual - data_email  # Calcula a diferença de tempo
##        return delta.days  # Retorna o número de dias passados
##    except Exception as e:
##        logging.error(f"Erro ao calcular dias passados: {str(e)}")
##        return 0  # Retorna 0 em caso de erro

def dias_passados(data_str):
    try:
        data_email = datetime.strptime(data_str, '%Y-%m-%d %H:%M:%S')  # Converte string de data para objeto datetime
        data_atual = datetime.now()  # Obtém a data e hora atuais
        delta = data_atual - data_email  # Calcula a diferença de tempo
        return delta.days  # Retorna o número de dias passados
    except Exception as e:
        logging.error(f"Erro ao calcular dias passados: {str(e)}")
        return 0  # Retorna 0 em caso de erro, evitando que o erro cause falha no processamento



def ajustar_colunas(sheet):
    colunas_necessarias = ['Email', 'Nome Completo', 'Cargo', 'Empresa', 'Primeiro E-MAIL ENVIADO?', 'Segundo email enviado?', 'terceiro email enviado?']
    header = sheet[1]
    print('o heather é:   ',header)
    print('o tipo do header é:  ',type(header))
    # Verificando as colunas existentes
    colunas_existentes = [cell.value for cell in header]
    novasColunas={}
    for index,value in enumerate(colunas_existentes):
        novasColunas[value]  =  index  

    colunas_existentes = novasColunas
    print("colunas existentes é:",colunas_existentes)
    # Se alguma coluna necessária não existir, criá-la
    tiveQueCriar = False
##    for coluna in colunas_necessarias:
##        if coluna not in colunas_existentes.keys():
##            sheet.cell(row=1, column=len(colunas_existentes) + 1, value=coluna)
##            colunas_existentes[coluna] = len(colunas_existentes) + 1
##            tiveQueCriar = True
##    
        

    return tiveQueCriar,colunas_existentes

# Função para obter o índice de uma coluna pelo nome
def obter_indice_coluna(colunas_existentes, nome_coluna):
    try:
        # Procurar a coluna pelo nome
        return colunas_existentes[nome_coluna] # Adiciona 1 para corresponder à indexação da planilha
    except ValueError:
        logging.error(f"Coluna '{nome_coluna}' não encontrada na planilha.")
        return None



# Função para enviar um email com tentativas adicionais
# Na função enviar_email
async def enviar_email(row, assunto, corpo, email, nome, cargo, empresa):
    global contador_emails_enviados  # Usando o contador global
    
    tentativas = 3  # Número máximo de tentativas
    for tentativa in range(tentativas):
        try:
            msg = MIMEMultipart()
            msg['From'] = FROM_EMAIL
            msg['To'] = email
            msg['Subject'] = assunto
            
            # Verificar e converter células para valores, garantindo que são strings
            corpo = str(corpovalue) if isinstance(corpo, openpyxl.cell.cell.Cell) else str(corpo or '')
            nome  = str(nome.value) if isinstance(nome, openpyxl.cell.cell.Cell) else str(nome or '')
            cargo = str(cargo.value) if isinstance(cargo, openpyxl.cell.cell.Cell) else str(cargo or '')
            empresa = str(empresa.value) if isinstance(empresa, openpyxl.cell.cell.Cell) else str(empresa or '')

            # Substituir variáveis no corpo do e-mail
            variaveis = {'funcionario': nome, 'cargo': cargo, 'empresa': empresa}
            for chave, valor in variaveis.items():
                corpo = corpo.replace(chave, valor)

            msg.attach(MIMEText(corpo, 'html'))

            # Conexão SMTP
            server = smtplib.SMTP(EMAIL_HOST, EMAIL_PORT)
            server.starttls()
            server.login(EMAIL_USER, EMAIL_PASSWORD)

            # Enviando e-mail
            server.sendmail(FROM_EMAIL, email, msg.as_string())
            server.quit()
            
            contador_emails_enviados += 1  # Incrementando o contador
            print(f'E-mail enviado com sucesso para {email}. E-mails enviados hoje: {contador_emails_enviados}/{LIMITE_DIARIO}')
            global emails_enviados
            emails_enviados.append(email)
            return True
        except smtplib.SMTPAuthenticationError as e:
            logging.error(f"Erro de autenticação ao enviar e-mail para {email}: {str(e)}")
            print(f'Erro de autenticação ao enviar e-mail para {email}')
            return 'Erro de autenticação'
        except smtplib.SMTPRecipientsRefused as e:
            logging.error(f"Destinatário recusado ao enviar e-mail para {email}: {str(e)}")
            print(f'Destinatário recusado ao enviar e-mail para {email}')
            return 'Destinatário recusado'
        except smtplib.SMTPConnectError as e:
            logging.error(f"Erro de conexão ao enviar e-mail para {email}: {str(e)}")
            if tentativa < tentativas - 1:  # Tentar novamente até o limite de tentativas
                print(f"Tentando novamente... {tentativa + 1}/{tentativas}")
                await asyncio.sleep(5)  # Espera de 5 segundos antes de tentar novamente
                continue
            else:
                logging.error(f"Falha ao conectar após {tentativas} tentativas para {email}.")
            return 'Erro de conexão'
        except Exception as e:
            logging.error(f"Erro inesperado ao enviar e-mail para {email}: {str(e)}")
            print(f'Erro inesperado ao enviar e-mail para {email}: {str(e)}')
            return 'Erro inesperado'


# Função para salvar a planilha com re-tentativas
async def salvar_planilha(workbook):
    tentativas_max = 5
    for tentativa in range(tentativas_max):
        try:
            workbook.save('emails_atualizados.xlsx')
            print("Planilha salva com sucesso.")
            return
        except Exception as e:
            logging.error(f"Erro ao salvar a planilha (Tentativa {tentativa + 1}): {str(e)}")
            print(f'Erro ao salvar a planilha: {str(e)}')
            if tentativa == tentativas_max - 1:
                logging.error(f"Falha ao salvar a planilha após {tentativas_max} tentativas.")
                print(f"Falha ao salvar a planilha após {tentativas_max} tentativas.")
                break
            await asyncio.sleep(30)  # Esperar 30 segundos antes de tentar novamente

# Função para enviar e-mails com limite de concorrência
async def enviar_email_com_concorrencia(row, assunto, corpo, email, nome, cargo, empresa, semaphore):
    # Usar o semaphore para garantir que no máximo 2 e-mails sejam enviados ao mesmo tempo
    async with semaphore:
        sucesso = await enviar_email(row, assunto, corpo, email, nome, cargo, empresa)
        return sucesso

def verificar_emails_enviados_no_dia():
    global contador_emails_enviados
    hoje = datetime.now().strftime('%Y-%m-%d')
    
    # Verificar no log se há e-mails enviados hoje
    try:
        with open(LOG_FILE, 'r') as log_file:
            for linha in log_file:
                if hoje in linha and "E-mail enviado com sucesso" in linha:
                    contador_emails_enviados += 1
    except FileNotFoundError:
        logging.info("Arquivo de log não encontrado. Contagem de e-mails iniciada.")
        
# Função para processar os e-mails
async def processar_emails():
    sheet, workbook = carregar_planilha()
##    ajustar_colunas(sheet)
    # Verifica se a planilha foi carregada corretamente
    if sheet is None or workbook is None:
        logging.error("Não foi possível carregar a planilha, o processo será encerrado.")
        print("Não foi possível carregar a planilha. Encerrando o programa.")
        return  # Não prosseguir se a planilha não foi carregada corretamente

    verificar_emails_enviados_no_dia()
    if contador_emails_enviados >= LIMITE_DIARIO:
        logging.info(f"Limite de {LIMITE_DIARIO} e-mails enviados no dia alcançado. Encerrando o programa.")
        print(f"Limite de {LIMITE_DIARIO} e-mails enviados no dia alcançado. Encerrando o programa.")
        return
    
    tiveQueCriar,colunas_existentes = ajustar_colunas(sheet)
    if tiveQueCriar:
        print('tive que criar coluna nova')
        sheet, workbook = carregar_planilha()
        if sheet is None or workbook is None:
            logging.error("Não foi possível carregar a planilha, o processo será encerrado.")
            print("Não foi possível carregar a planilha. Encerrando o programa.")
            return  # Não prosseguir se a planilha não foi carregada corretamente
    

        
    row_index = 2  # Começando da segunda linha para evitar o cabeçalho

    semaphore = asyncio.Semaphore(2)  # Limita para 2 e-mails sendo enviados ao mesmo tempo

    # Loop para percorrer as linhas da planilha
    while row_index <= sheet.max_row:
        try:
            if contador_emails_enviados >= LIMITE_DIARIO:
                logging.info(f"Limite de {LIMITE_DIARIO} e-mails enviados no dia alcançado. Encerrando o programa.")
                print(f"Limite de {LIMITE_DIARIO} e-mails enviados no dia alcançado. Encerrando o programa.")
                break  # Interrompe o loop se o limite de e-mails foi alcançado
            
            row = sheet[row_index]
            row2=[]
            for x in row:
                row2.append(x.value)

            row=copy.copy(row2)
##            print(row)
##            try:
            nome = row[colunas_existentes['Nome Completo']]
##            except:
##                print('foi no nome que deu erro')
##            try:
            cargo = row[colunas_existentes['CARGO']]
            
##            except:
####                print('foi no cargo')
##            try:
            empresa = row[colunas_existentes['RAZÃO SOCIAL']] or row[colunas_existentes['NOME FANTASIA']] or None
##            except:
##                print('foi na razao social ou no nome fantasia')
##            try:
            email = row[colunas_existentes['EMAIL']]
##            except:
##                print('deu erro no email')
            global emails_enviados
            if email in emails_enviados:
                continue
            corpo_primeiro_email = [x.value for x in sheet[2]][colunas_existentes['Corpo primeiro e-mail']]
            print("colunas_existentes['Corpo primeiro e-mail'] é ",colunas_existentes['Corpo primeiro e-mail'])
            
            if corpo_primeiro_email is None:
                corpo_primeiro_email = "Mensagem padrão caso o corpo não esteja definido."
            else:
                corpo_primeiro_email = str(corpo_primeiro_email)
                
            corpo_segundo_email = [x.value for x in sheet[2]][colunas_existentes['Corpo segundo e-mail']]
            if corpo_segundo_email is None:
                corpo_segundo_email = "Mensagem padrão caso o corpo não esteja definido."
            else:
                corpo_segundo_email = str(corpo_segundo_email)

            indice=colunas_existentes['Terceiro email enviado?']
            corpo_terceiro_email = [x.value for x in sheet[2]][indice]
            if corpo_terceiro_email is None:
                corpo_terceiro_email = "Mensagem padrão caso o corpo não esteja definido."
            else:
                corpo_terceiro_email = str(corpo_terceiro_email)
                
            corpo_primeiro_email = str(corpo_primeiro_email ) if isinstance(corpo_primeiro_email, openpyxl.cell.cell.Cell) else str(corpo_primeiro_email or '')
            corpo_segundo_email = str(corpo_segundo_email ) if isinstance(corpo_segundo_email, openpyxl.cell.cell.Cell) else str(corpo_segundo_email or '')
            corpo_terceiro_email = str(corpo_terceiro_email ) if isinstance(corpo_terceiro_email, openpyxl.cell.cell.Cell) else str(corpo_terceiro_email or '')
            nome = str(nome ) if isinstance(nome, openpyxl.cell.cell.Cell) else str(nome or '')
            cargo = str(cargo ) if isinstance(cargo, openpyxl.cell.cell.Cell) else str(cargo or '')
            empresa = str(empresa ) if isinstance(empresa, openpyxl.cell.cell.Cell) else str(empresa or '')

##            indice_primeiro_email  = obter_indice_coluna(colunas_existentes, 'Primeiro E-MAIL ENVIADO?')
            print('o colunas existentes logo antes do erro é:  ',colunas_existentes)
            indice_primeiro_email  = colunas_existentes['Primeiro E-MAIL ENVIADO?']
            indice_segundo_email   = obter_indice_coluna(colunas_existentes, 'Segundo email enviado?')
            indice_terceiro_email  = obter_indice_coluna(colunas_existentes, 'Terceiro email enviado?')
            print('a row é:',row)
            print('o indice do primeiro email é:',indice_segundo_email)
            primeiro_email = row[indice_primeiro_email]
            print('o primeiro email é: ', primeiro_email)
            segundo_email  = row[indice_segundo_email] 
            terceiro_email = row[indice_terceiro_email] 
            print('o email é:', email)
            # Verifica se o e-mail ainda não foi enviado
            if email and primeiro_email not in ['Email inválido', None]:
                print('o email é:', email)
            
                # Verificar e converter células para valores
##                nome = str(nome ) if isinstance(nome, openpyxl.cell.cell.Cell) else str(nome or '')
##                cargo = str(cargo ) if isinstance(cargo, openpyxl.cell.cell.Cell) else str(cargo or '')
##                empresa = str(empresa ) if isinstance(empresa, openpyxl.cell.cell.Cell) else str(empresa or '')
##                
                resultado = await enviar_email_com_concorrencia(row, 'Benefício Gratuito para sua Equipe', corpo_primeiro_email, email, nome, cargo, empresa, semaphore)
                if resultado == 'Erro de autenticação':
                    logging.error(f'Falha na autenticação ao tentar enviar para {email}')
                    break  # Se falhar na autenticação, para o processo
                elif resultado == 'Destinatário recusado':
                    # Marcar como inválido
                    row[colunas_existentes['Primeiro E-MAIL ENVIADO?']]  = 'Email inválido'
                elif resultado == 'Erro inesperado':
                    # Manter tentativa de envio
                    pass
                else:
##                    indiceDoEmail  =  colunas_existentes['Primeiro EMAIL ENVIADO?']
                    altera_tabela(row_index,indice_primeiro_email,time.strftime('%Y-%m-%d %H:%M:%S'),sheet)
                    row[colunas_existentes['Primeiro E-MAIL ENVIADO?']]  = time.strftime('%Y-%m-%d %H:%M:%S')
##                    sheet.cell(row=row_index, column=indice_primeiro_email, value=time.strftime('%Y-%m-%d %H:%M:%S'))
                    #sheet[row_index][indiceDoEmail]  = time.strftime('%Y-%m-%d %H:%M:%S')
            

            # Verifica se já passou o tempo para enviar o segundo e-mail
            print('o valor do primeiro_email é: ',primeiro_email)
            if segundo_email is None and primeiro_email and dias_passados(primeiro_email) >= 5:
                corpo_segundo_email = row[colunas_existentes['Corpo segundo e-mail']] 
                if corpo_segundo_email is None:
                    corpo_segundo_email = "Mensagem padrão para o segundo e-mail."
                else:
                    corpo_segundo_email = str(corpo_segundo_email)
                    
                resultado = await enviar_email_com_concorrencia(row, 'Assunto para o segundo e-mail', corpo_segundo_email, email, nome, cargo, empresa, semaphore)
                if resultado == 'Erro de autenticação':
                    logging.error(f'Falha na autenticação ao tentar enviar para {email}')
                    break
                elif resultado == 'Destinatário recusado':
                    row[colunas_existentes['Segundo email enviado?']]  = 'Email inválido'
                elif resultado == 'Erro inesperado':
                    pass
                else:
                    altera_tabela(row_index,indice_segundo_email,time.strftime('%Y-%m-%d %H:%M:%S'),sheet)
##                    sheet.cell(row=row_index, column=indice_segundo_email, value=time.strftime('%Y-%m-%d %H:%M:%S'))
                    row[colunas_existentes['Segundo email enviado?']]  = time.strftime('%Y-%m-%d %H:%M:%S')

            # Verifica se já passou o tempo para enviar o terceiro e-mail
            print('o valor do segundo_email é: ',segundo_email)
            if terceiro_email is None and segundo_email and dias_passados(segundo_email) >= 7:
                corpo_terceiro_email = row[colunas_existentes['Corpo terceiro e-mail']] 
                if corpo_terceiro_email is None:
                    corpo_terceiro_email = "Mensagem padrão para o terceiro e-mail."
                else:
                    corpo_terceiro_email = str(corpo_terceiro_email)
                
                resultado = await enviar_email_com_concorrencia(row, 'Assunto para o terceiro e-mail', corpo_terceiro_email, email, nome, cargo, empresa, semaphore)
                if resultado == 'Erro de autenticação':
                    logging.error(f'Falha na autenticação ao tentar enviar para {email}')
                    break
                elif resultado == 'Destinatário recusado':
                    row[colunas_existentes['terceiro email enviado?']]  = 'Email inválido'
                elif resultado == 'Erro inesperado':
                    pass
                else:
                    altera_tabela(row_index,indice_terceiro_email,time.strftime('%Y-%m-%d %H:%M:%S'),sheet)
##                    sheet.cell(row=row_index, column=indice_terceiro_email, value=time.strftime('%Y-%m-%d %H:%M:%S'))
                    row[colunas_existentes['Terceiro email enviado?']]  = time.strftime('%Y-%m-%d %H:%M:%S')

            row_index += 1

            # Salvar a planilha a cada envio de e-mail
            global alterou_tabela
            if alterou_tabela:
                await salvar_planilha(workbook)
                alterou_tabela = False

            # Esperar antes de enviar o próximo e-mail
            intervalo = random.uniform(intervalo_min, intervalo_max)
            await asyncio.sleep(intervalo)

        except Exception as e:
            logging.error(f"Erro ao processar a linha {row_index}: {str(e)}")
            print(f"Erro ao processar a linha {row_index}. Erro: {str(e)}")
            row_index += 1  # Continuação do processamento, mas registrando o erro
    print('Processamento de e-mails concluído!')


# Função principal para rodar o código
async def main():
    await processar_emails()

if __name__ == '__main__':
    asyncio.run(main())
